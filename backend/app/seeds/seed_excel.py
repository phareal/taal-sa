"""Charge le fichier `DASHBOARD TAAL-SA.xlsx` dans la base PostgreSQL.

Règles :
- Idempotent : purge les 6 tables (ordre FK) puis ré-insère tout.
- Schéma créé automatiquement si absent (SQLAlchemy metadata.create_all).
- Les navires, shippers et clients sont dérivés des lignes de connaissements
  puis enrichis par les feuilles de synthèse (CA, statut, nb BL…).

Utilisation :
    python -m app.seeds.seed_excel [/chemin/vers/fichier.xlsx]

Par défaut : /app/data/seed/DASHBOARD-TAAL-SA.xlsx
"""

from __future__ import annotations

import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.database import Base, SyncSessionLocal, sync_engine
from app.models import Client, Connaissement, Cotation, Navire, Prospect, Shipper

DEFAULT_XLSX = Path("/app/data/seed/DASHBOARD-TAAL-SA.xlsx")

SHEET_CONSOLIDEES = "📋 DONNÉES CONSOLIDÉES"
SHEET_2026 = "📈 2026 (MARGES)"
SHEET_SHIPPERS = "🚚 SHIPPERS À RELANCER"
SHEET_CLIENTS = "👥 SYNTHÈSE CLIENTS"
SHEET_PROSPECTS = "🎯 SYNTHÈSE PROSPECTS"
SHEET_COTATIONS = "💼 Opportunités"

MOIS_FR = {
    "JANVIER": 1, "FEVRIER": 2, "FÉVRIER": 2, "MARS": 3, "AVRIL": 4,
    "MAI": 5, "JUIN": 6, "JUILLET": 7, "AOUT": 8, "AOÛT": 8,
    "SEPTEMBRE": 9, "OCTOBRE": 10, "NOVEMBRE": 11, "DECEMBRE": 12, "DÉCEMBRE": 12,
}


# ─────────────────────── helpers ───────────────────────


def log(msg: str) -> None:
    print(f"[seed] {msg}", flush=True)


def clean_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def clean_key(value: Any) -> str | None:
    """Clé normalisée pour dédoublonnage (trim + uppercase, diacritiques conservés)."""
    s = clean_str(value)
    return s.upper() if s else None


def to_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return None


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def normalize_statut(value: Any) -> str:
    s = clean_str(value)
    if not s:
        return "ACTIF"
    up = s.upper()
    if "PERDU" in up:
        return "PERDU"
    if "DÉCLIN" in up or "DECLIN" in up:
        return "EN_DÉCLIN"
    return "ACTIF"


def normalize_resultat(value: Any) -> str:
    s = clean_str(value)
    if not s:
        return "EN COURS"
    up = s.upper()
    if "GAGN" in up:
        return "GAGNÉ"
    if "PERDU" in up:
        return "PERDU"
    if "ANNUL" in up:
        return "ANNULÉ"
    return "EN COURS"


def iter_data_rows(ws, start_row: int) -> Iterable[tuple]:
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if any(c is not None and str(c).strip() != "" for c in row):
            yield row


# ─────────────────────── seeding par table ───────────────────────


def reset_tables(db: Session) -> None:
    log("purge des tables existantes…")
    db.query(Cotation).delete()
    db.query(Prospect).delete()
    db.query(Connaissement).delete()
    db.query(Client).delete()
    db.query(Shipper).delete()
    db.query(Navire).delete()
    db.commit()


def seed_from_connaissements(db: Session, xlsx_path: Path) -> dict[str, dict[str, int]]:
    """Parcourt les feuilles de connaissements, crée navires/shippers/clients
    et retourne les maps de clé normalisée → id."""

    log(f"lecture du fichier Excel : {xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)

    navires: dict[str, str] = {}  # key → nom original
    shippers: dict[str, str] = {}
    clients: dict[str, str] = {}
    bls: list[dict] = []

    # Feuille 2019–2025 ─────────────────────────────────────
    ws = wb[SHEET_CONSOLIDEES]
    for row in iter_data_rows(ws, start_row=6):
        # (N°, NAVIRE, BL, SHIPPER, CONSIG, QTE, POIDS, VOL, DOCS_FEES, MOIS, ANNEE)
        navire = clean_str(row[1])
        numero_bl = clean_str(row[2])
        shipper = clean_str(row[3])
        client = clean_str(row[4])
        quantite = clean_str(row[5])
        poids = to_float(row[6])
        volume = to_float(row[7])
        docs_fees = to_int(row[8])
        mois = clean_str(row[9])
        annee = to_int(row[10])

        if annee is None:
            continue

        if navire:
            navires.setdefault(clean_key(navire), navire)
        if shipper:
            shippers.setdefault(clean_key(shipper), shipper)
        if client:
            clients.setdefault(clean_key(client), client)

        bls.append({
            "numero_bl": numero_bl,
            "navire_key": clean_key(navire),
            "shipper_key": clean_key(shipper),
            "client_key": clean_key(client),
            "annee": annee,
            "mois": mois,
            "mois_num": MOIS_FR.get(mois.upper()) if mois else None,
            "quantite": quantite,
            "poids_kg": poids,
            "volume_m3": volume,
            "docs_fees_fcfa": docs_fees,
            "montant_normal_fcfa": None,
            "marge_fcfa": None,
            "taux_marge": None,
        })

    # Feuille 2026 avec marges ─────────────────────────────
    ws = wb[SHEET_2026]
    for row in iter_data_rows(ws, start_row=9):
        # (N°, NAVIRE, BL, SHIPPER, CONSIG, QTE, POIDS, VOL, DOCS_FEES, MONTANT_NORMAL, MARGE, TAUX)
        navire = clean_str(row[1])
        numero_bl = clean_str(row[2])
        shipper = clean_str(row[3])
        client = clean_str(row[4])
        quantite = clean_str(row[5])
        poids = to_float(row[6])
        volume = to_float(row[7])
        docs_fees = to_int(row[8])
        montant_normal = to_int(row[9])
        marge = to_int(row[10])
        taux_marge = to_float(row[11])

        if navire:
            navires.setdefault(clean_key(navire), navire)
        if shipper:
            shippers.setdefault(clean_key(shipper), shipper)
        if client:
            clients.setdefault(clean_key(client), client)

        bls.append({
            "numero_bl": numero_bl,
            "navire_key": clean_key(navire),
            "shipper_key": clean_key(shipper),
            "client_key": clean_key(client),
            "annee": 2026,
            "mois": None,
            "mois_num": None,
            "quantite": quantite,
            "poids_kg": poids,
            "volume_m3": volume,
            "docs_fees_fcfa": docs_fees,
            "montant_normal_fcfa": montant_normal,
            "marge_fcfa": marge,
            "taux_marge": taux_marge,
        })

    wb.close()
    log(f"  → {len(navires)} navires, {len(shippers)} shippers, {len(clients)} clients, {len(bls)} BLs")

    # Insertion navires / shippers / clients ───────────────
    navire_id_by_key: dict[str, int] = {}
    for key, nom in navires.items():
        obj = Navire(nom=nom, actif=True)
        db.add(obj)
        db.flush()
        navire_id_by_key[key] = obj.id

    shipper_id_by_key: dict[str, int] = {}
    for key, nom in shippers.items():
        obj = Shipper(nom=nom, statut="ACTIF")
        db.add(obj)
        db.flush()
        shipper_id_by_key[key] = obj.id

    client_id_by_key: dict[str, int] = {}
    for key, nom in clients.items():
        obj = Client(nom=nom)
        db.add(obj)
        db.flush()
        client_id_by_key[key] = obj.id

    db.commit()
    log("  → navires/shippers/clients insérés")

    # Insertion connaissements ─────────────────────────────
    log(f"insertion de {len(bls)} connaissements…")
    batch: list[Connaissement] = []
    for bl in bls:
        batch.append(Connaissement(
            numero_bl=bl["numero_bl"],
            navire_id=navire_id_by_key.get(bl["navire_key"]) if bl["navire_key"] else None,
            shipper_id=shipper_id_by_key.get(bl["shipper_key"]) if bl["shipper_key"] else None,
            client_id=client_id_by_key.get(bl["client_key"]) if bl["client_key"] else None,
            annee=bl["annee"],
            mois=bl["mois"],
            mois_num=bl["mois_num"],
            quantite=bl["quantite"],
            poids_kg=bl["poids_kg"],
            volume_m3=bl["volume_m3"],
            docs_fees_fcfa=bl["docs_fees_fcfa"],
            montant_normal_fcfa=bl["montant_normal_fcfa"],
            marge_fcfa=bl["marge_fcfa"],
            taux_marge=bl["taux_marge"],
            type_operation="IMPORT",
        ))
        if len(batch) >= 500:
            db.bulk_save_objects(batch)
            db.commit()
            batch = []
    if batch:
        db.bulk_save_objects(batch)
        db.commit()

    return {
        "navires": navire_id_by_key,
        "shippers": shipper_id_by_key,
        "clients": client_id_by_key,
    }


def enrich_shippers(db: Session, xlsx_path: Path, maps: dict) -> int:
    """Met à jour CA / nb BL / statut pour les shippers listés dans la feuille de relance."""
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[SHEET_SHIPPERS]

    updated = 0
    for row in iter_data_rows(ws, start_row=6):
        nom = clean_str(row[0])
        if not nom:
            continue
        key = clean_key(nom)
        shipper_id = maps["shippers"].get(key)
        if shipper_id is None:
            # créer le shipper s'il n'a pas de BL mais figure dans la liste de relance
            obj = Shipper(nom=nom)
            db.add(obj)
            db.flush()
            maps["shippers"][key] = obj.id
            shipper_id = obj.id

        db.query(Shipper).filter(Shipper.id == shipper_id).update({
            "ca_passe_fcfa": to_int(row[1]) or 0,
            "ca_actuel_fcfa": to_int(row[2]) or 0,
            "nb_bl_passe": to_int(row[5]) or 0,
            "nb_bl_actuel": to_int(row[6]) or 0,
            "statut": normalize_statut(row[7]),
        })
        updated += 1

    wb.close()
    db.commit()
    return updated


def seed_prospects(db: Session, xlsx_path: Path, maps: dict) -> int:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[SHEET_PROSPECTS]

    inserted = 0
    for row in iter_data_rows(ws, start_row=17):
        ordre = to_int(row[0])
        nom_client = clean_str(row[1])
        if not nom_client:
            continue

        key = clean_key(nom_client)
        client_id = maps["clients"].get(key)
        if client_id is None:
            obj = Client(nom=nom_client)
            db.add(obj)
            db.flush()
            maps["clients"][key] = obj.id
            client_id = obj.id

        # priorité : top 5 = P1, 6-10 = P2, reste = P3
        if ordre is None:
            priorite = "P3"
        elif ordre <= 5:
            priorite = "P1"
        elif ordre <= 10:
            priorite = "P2"
        else:
            priorite = "P3"

        ca_pic = to_int(row[2]) or 0
        ca_derniere = to_int(row[3]) or 0
        annee_derniere = to_int(row[7])
        statut = clean_str(row[8])
        action = clean_str(row[9])

        db.add(Prospect(
            client_id=client_id,
            type_statut=statut,
            priorite=priorite,
            ca_pic_fcfa=ca_pic,
            ca_derniere_fcfa=ca_derniere,
            annee_derniere_op=annee_derniere,
            action_prevue=action,
            statut_relance="À CONTACTER",
        ))
        inserted += 1

    wb.close()
    db.commit()
    return inserted


def seed_cotations(db: Session, xlsx_path: Path, maps: dict) -> int:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[SHEET_COTATIONS]

    inserted = 0
    for row in iter_data_rows(ws, start_row=16):
        numero = clean_str(row[0])
        if not numero:
            continue

        raw_date = row[1]
        date_cotation: date | None = None
        if isinstance(raw_date, datetime):
            date_cotation = raw_date.date()
        elif isinstance(raw_date, date):
            date_cotation = raw_date

        nom_client = clean_str(row[2])
        client_id = None
        if nom_client:
            key = clean_key(nom_client)
            client_id = maps["clients"].get(key)
            if client_id is None:
                obj = Client(nom=nom_client)
                db.add(obj)
                db.flush()
                maps["clients"][key] = obj.id
                client_id = obj.id

        type_service = clean_str(row[3])
        offre = to_float(row[4])
        cotation_client = to_float(row[5])
        marge = to_float(row[6])
        resultat = normalize_resultat(row[7])
        observations = clean_str(row[8])

        db.add(Cotation(
            numero_cotation=numero,
            date_cotation=date_cotation,
            client_id=client_id,
            type_service=type_service,
            offre_transitaire=offre,
            cotation_client=cotation_client,
            marge=marge,
            devise="EUR",
            resultat=resultat,
            observations=observations,
        ))
        inserted += 1

    wb.close()
    db.commit()
    return inserted


# ─────────────────────── entrypoint ───────────────────────


def run(xlsx_path: Path) -> None:
    if not xlsx_path.exists():
        log(f"ERREUR : fichier introuvable → {xlsx_path}")
        sys.exit(1)

    log("création des tables si nécessaire…")
    Base.metadata.create_all(sync_engine)

    with SyncSessionLocal() as db:
        reset_tables(db)
        maps = seed_from_connaissements(db, xlsx_path)

        n_shippers = enrich_shippers(db, xlsx_path, maps)
        log(f"  → {n_shippers} shippers enrichis (CA / statut)")

        n_prospects = seed_prospects(db, xlsx_path, maps)
        log(f"  → {n_prospects} prospects insérés")

        n_cotations = seed_cotations(db, xlsx_path, maps)
        log(f"  → {n_cotations} cotations insérées")

    log("✔ seed terminé")


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    run(xlsx)


if __name__ == "__main__":
    main()
